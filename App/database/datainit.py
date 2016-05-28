#!/usr/bin/env python
#-*-coding:utf8-*-

from openpyxl import load_workbook
import re
import pydb

path = "book_top250.xlsx"

wb = load_workbook(path)
ws = wb.active

db = pydb.mydb()

index = 2
while True:
	isbn = ws.cell(row = index,column = 1).value
	title = ws.cell(row = index,column = 2).value
	author = ws.cell(row = index,column = 3).value
	Press = ws.cell(row = index,column = 4).value
	date = ws.cell(row = index,column = 5).value
	pagenum = ws.cell(row = index,column = 6).value
	tags = ws.cell(row = index,column = 7).value

	if not isbn:
		break
	#处理出版年月
	pattern = re.compile(u"[-年/]")
	s = pattern.search(date)
	if s:
		date = date[:s.start()].strip()
	#处理页数
	i = pagenum.find(u'页')
	if i != -1:
		pagenum = pagenum[:i]
	i = pagenum.find(u'；')
	if i != -1:
		pagenum = pagenum[:i].strip()

	db.insert((isbn,title,author,date,Press,pagenum,tags))
	index += 1

db.commit()
