#!/usr/bin/env python
#-*-coding:utf8-*-

import urllib2
from openpyxl import load_workbook,Workbook
from bs4 import BeautifulSoup
import re,random

indexurl = "https://book.douban.com/top250?start=0"

bookurl = None
ifend = False

hds = [{'User-Agent':'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50'},
{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50'},
{'User-Agent':'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0;'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)'},
{'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1'},
{'User-Agent':'Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1'},
{'User-Agent':'Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11'},
{'User-Agent':'Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11'},
{'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0)'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; TencentTraveler 4.0)'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; The World)'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SE 2.X MetaSr 1.0; SE 2.X MetaSr 1.0; .NET CLR 2.0.50727; SE 2.X MetaSr 1.0)'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; 360SE)'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Avant Browser)'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)'},
{'User-Agent':'Mozilla/5.0 (iPhone; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5'},
{'User-Agent':'Mozilla/5.0 (iPod; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5'},
{'User-Agent':'Mozilla/5.0 (iPad; U; CPU OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5'},
{'User-Agent':'Mozilla/5.0 (Linux; U; Android 2.3.7; en-us; Nexus One Build/FRF91) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1'},
{'User-Agent':'MQQBrowser/26 Mozilla/5.0 (Linux; U; Android 2.3.7; zh-cn; MB200 Build/GRJ22; CyanogenMod-7) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1'},
{'User-Agent':'Opera/9.80 (Android 2.3.4; Linux; Opera Mobi/build-1107180945; U; en-GB) Presto/2.8.149 Version/11.10'},
{'User-Agent':'Mozilla/5.0 (Linux; U; Android 3.0; en-us; Xoom Build/HRI39) AppleWebKit/534.13 (KHTML, like Gecko) Version/4.0 Safari/534.13'},
{'User-Agent':'Mozilla/5.0 (BlackBerry; U; BlackBerry 9800; en) AppleWebKit/534.1+ (KHTML, like Gecko) Version/6.0.0.337 Mobile Safari/534.1+'},
{'User-Agent':'Mozilla/5.0 (hp-tablet; Linux; hpwOS/3.0.0; U; en-US) AppleWebKit/534.6 (KHTML, like Gecko) wOSBrowser/233.70 Safari/534.6 TouchPad/1.0'},
{'User-Agent':'Mozilla/5.0 (SymbianOS/9.4; Series60/5.0 NokiaN97-1/20.0.019; Profile/MIDP-2.1 Configuration/CLDC-1.1) AppleWebKit/525 (KHTML, like Gecko) BrowserNG/7.1.18124'},
{'User-Agent':'Mozilla/5.0 (compatible; MSIE 9.0; Windows Phone OS 7.5; Trident/5.0; IEMobile/9.0; HTC; Titan)'},
{'User-Agent':'UCWEB7.0.2.37/28/999'},
{'User-Agent':'NOKIA5700/ UCWEB7.0.2.37/28/999'},
{'User-Agent':'Openwave/ UCWEB7.0.2.37/28/999'},
{'User-Agent':'Mozilla/4.0 (compatible; MSIE 6.0; ) Opera/UCWEB7.0.2.37/28/999'}
]

def  getHtml(url):
	 i = random.randint(0,len(hds)-1)
	 request = urllib2.Request(url)
	 request.add_header('User-Agent',hds[i]['User-Agent'])
	 response = urllib2.urlopen(request)
	 return response.read()

wb = Workbook()
ws = wb.active
ws['A1'] = "ISBN"
ws['B1'] = "书名"
ws['C1'] = "作者"
ws['D1'] = "出版社"
ws['E1'] ="出版年"
ws['F1'] = '页数'
ws['G1'] = '标签'

rowl = [2]

def writeBook(title,author,pcompany,pdate,pagenum,isbn,tags):
	row = rowl[0]
	ws.cell(row=row, column = 1).value = isbn
	ws.cell(row=row, column = 2).value = title
	ws.cell(row=row, column = 3).value = author
	ws.cell(row=row, column = 4).value = pcompany
	ws.cell(row=row, column = 5).value = pdate
	ws.cell(row=row, column = 6).value = pagenum
	ws.cell(row=row, column = 7).value = tags
	rowl[0] += 1
	print "success ", row - 1

errorlog = open('errorlog.txt','w+')

while not ifend:
	html = getHtml(indexurl)
	soup = BeautifulSoup(html,"html5lib")
	blist = soup.find('div',class_='indent').find_all('table')
	#print indexurl, len(blist)
	nextpagetag = soup.find('div',class_="paginator").find('a',text=re.compile(u'后页'))
	if nextpagetag == None:
		ifend = True
	else:
		indexurl = nextpagetag['href']
	num = 0

	for tag in blist:
		try:
			bookurl = tag.find('div',class_='pl2').find('a')['href']
			html = getHtml(bookurl)
			soup = BeautifulSoup(html,"html5lib")
			#获取书名
			title = soup.find('div',{'id':"wrapper"}).find('h1').text.strip()
			dtag = soup.find('div',class_='subject clearfix').find('div',{'id':'info'})
			#获取作者名
			author = dtag.find('span',text=re.compile(u'作者')).next_sibling.next_sibling.string.strip()
			#获取出版社
			pcompany = dtag.find('span',text=re.compile(u'出版社')).next_sibling.string.strip()
			#获取出版年月日
			pdate = dtag.find('span',text=re.compile(u'出版年')).next_sibling.string.strip()
			#获取页数
			pagenum = dtag.find('span',text=re.compile(u'页数')).next_sibling.string.strip()
			#ISBN
			isbn = dtag.find('span',text=re.compile(u'ISBN')).next_sibling.string.strip()
			#标签
			slist = soup.find('div',{'id':'db-tags-section'}).find('div',class_='indent').find_all('span')
			slist = slist[:5] if len(slist) > 5 else slist
			taglist = [x.text.strip() for x in slist]
			tags=";".join(taglist)
			print num,
			num += 1
			writeBook(title,author,pcompany,pdate,pagenum,isbn,tags)
		except Exception,e:
			errorlog.write(bookurl+':' + str(e) + "\n")
			continue

wb.save("top250.xlsx")
errorlog.close()
